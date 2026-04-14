#!/usr/bin/env python3
"""
generar_excel.py — ICB Prospecta
Genera el Excel Top 15 SKUs para un cliente gastronómico.

Uso:
    python generar_excel.py \
        --cliente "Sienna Bakery" \
        --tipo "PASTELERÍA" \
        --catalogo "ruta/al/catalogo.xlsx" \
        --familias "COBERTURA,HARINAS Y SEMOLAS,..." \
        --output "outputs/SiennaBakery_Top15_ICB.xlsx"

También se puede invocar como módulo pasando un dict de configuración:
    from scripts.generar_excel import generar_top15
    resultado = generar_top15(config)
"""
import argparse
import sys
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Helpers de estilo ────────────────────────────────────────────────────────

def fill(hex_c: str):
    return PatternFill("solid", fgColor=hex_c)

def border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def fmt(cell, value, bold=False, size=9, color="000000", bg=None,
        halign="left", valign="center", wrap=False):
    cell.value = value
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    if bg:
        cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = border()


# ── Colores ──────────────────────────────────────────────────────────────────

C_AZUL     = "1F4E79"
C_AZUL_MED = "2E75B6"
C_VERDE    = "1E6E3A"
C_VERDE_CLR= "D5EDDF"
C_AMARILLO = "FFF2CC"
C_GRIS_CLR = "F2F2F2"
C_BLANCO   = "FFFFFF"
C_NARANJA  = "E8711A"

PRIO_BG  = {"1": C_VERDE_CLR, "2": C_AMARILLO, "3": "FCE4EC"}
PRIO_LBL = {"1": "P1 — CORE", "2": "P2 — ALTO", "3": "P3 — COMPL."}


# ── Lectura del catálogo ─────────────────────────────────────────────────────

CATALOGO_URL = "https://1drv.ms/x/c/91cea5c809fe1373/IQAoUK8ItuU8R4k0BCHZhvIBAZubGKc0-FvXZuOJ-2-296o?e=JAf6qD&download=1"

def descargar_catalogo(destino: str = "/tmp/icb_catalogo.xlsx") -> str:
    """Descarga el catálogo ICB desde OneDrive. Fallback a archivos locales."""
    import urllib.request, glob
    try:
        urllib.request.urlretrieve(CATALOGO_URL, destino)
        return destino
    except Exception:
        pass
    for patron in [
        "/sessions/*/mnt/uploads/*.xlsx",
        "/sessions/*/mnt/outputs/*.xlsx",
    ]:
        matches = [f for f in glob.glob(patron)
                   if any(k in f for k in ("LP_", "Lista Precios", "lista_precios", "catalogo"))]
        if matches:
            return sorted(matches)[-1]
    raise FileNotFoundError("No se encontró el catálogo ICB. Adjunta el archivo XLSX o verifica el link de OneDrive.")

def leer_catalogo(ruta: str = None) -> pd.DataFrame:
    if ruta is None:
        ruta = descargar_catalogo()
    xl = pd.ExcelFile(ruta)
    frames = []
    for sheet in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet)
        df["BASE"] = sheet.upper()
        frames.append(df)
    df = pd.concat(frames, ignore_index=True)
    df["CÓDIGO"] = df["CÓDIGO"].astype(str)
    df["PRECIO"]    = pd.to_numeric(df.get("PRECIO"),    errors="coerce").fillna(0)
    df["PRECIO UN"] = pd.to_numeric(df.get("PRECIO UN"), errors="coerce").fillna(0)
    return df


# ── Selección de Top 15 ──────────────────────────────────────────────────────

def seleccionar_top15(df: pd.DataFrame, familias_ordenadas: list,
                      skus_config: list = None) -> pd.DataFrame:
    """
    familias_ordenadas: lista de dicts con keys:
        familia (str), prioridad (str: '1'|'2'|'3'), consumo (str), motivo (str)
    skus_config: lista opcional de dicts con codigo_preferido por familia
    """
    resultados = []
    codigos_usados = set()

    for item in familias_ordenadas:
        fam = item["familia"]
        subset = df[df["FAMILIA"] == fam].copy()
        if subset.empty:
            continue

        # Preferir PRODUCTO FOCO = SI, luego precio más alto, evitar liquidación si hay nacional
        nacional = subset[subset["BASE"] == "NACIONAL"]
        pool = nacional if not nacional.empty else subset

        # Si hay código preferido explícito
        codigo_pref = item.get("codigo_preferido")
        if codigo_pref:
            exact = pool[pool["CÓDIGO"] == str(codigo_pref)]
            if not exact.empty:
                prod = exact.iloc[0]
                if prod["CÓDIGO"] not in codigos_usados:
                    resultados.append(_build_row(prod, item))
                    codigos_usados.add(prod["CÓDIGO"])
                    continue

        # Ordenar: foco primero, luego precio desc
        pool = pool.copy()
        pool["_foco"] = (pool.get("PRODUCTO FOCO", "").astype(str).str.upper() == "SI").astype(int)
        pool = pool.sort_values(["_foco", "PRECIO"], ascending=[False, False])

        for _, prod in pool.iterrows():
            if prod["CÓDIGO"] not in codigos_usados:
                resultados.append(_build_row(prod, item))
                codigos_usados.add(prod["CÓDIGO"])
                break

        if len(resultados) >= 15:
            break

    return resultados[:15]


def _build_row(prod, item: dict) -> dict:
    precio = float(prod.get("PRECIO", 0) or 0)
    precio_un = float(prod.get("PRECIO UN", 0) or 0)

    # Estimación consumo mensual según prioridad y tamaño asumido (1 local nuevo)
    freq = {"1": 4, "2": 2, "3": 1}.get(str(item["prioridad"]), 1)
    consumo_mensual = precio * freq

    return {
        "prioridad":  str(item["prioridad"]),
        "familia":    item["familia"],
        "codigo":     str(prod["CÓDIGO"]),
        "descripcion": str(prod["DESCRIPCIÓN"]),
        "umv":        str(prod.get("DESCRIPCION UMV", "")),
        "precio":     precio,
        "precio_un":  precio_un,
        "consumo_mensual": consumo_mensual,
        "motivo":     item.get("motivo", ""),
        "base":       str(prod.get("BASE", "")),
        "foco":       "★" if str(prod.get("PRODUCTO FOCO","")).upper() == "SI" else "",
    }


# ── Generación del Excel ─────────────────────────────────────────────────────

def generar_excel(resultados: list, nombre_cliente: str, tipo_cliente: str,
                  output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top 15 SKUs"

    hoy = datetime.today().strftime("%B %Y").upper()

    # ── Título ──
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = f"TOP 15 SKUs ICB — {nombre_cliente.upper()}"
    c.font  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill  = fill(C_AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = f"{tipo_cliente}  •  Lista ICB {hoy}  •  Generado: {datetime.today().strftime('%d/%m/%Y')}"
    c.font  = Font(name="Arial", italic=True, size=9, color="595959")
    c.fill  = fill("EBF3FB")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 6

    # ── Encabezados ──
    hdrs = ["#", "PRIORIDAD", "FAMILIA", "CÓDIGO", "DESCRIPCIÓN",
            "UMV", "$/CAJA", "$/UN", "POR QUÉ PARA ESTE CLIENTE"]
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=4, column=ci)
        fmt(c, h, bold=True, color="FFFFFF", bg=C_AZUL_MED, halign="center")
    ws.row_dimensions[4].height = 22

    # ── Filas de datos ──
    for ri, row in enumerate(resultados, 5):
        prio   = row["prioridad"]
        row_bg = C_GRIS_CLR if ri % 2 == 0 else C_BLANCO
        prio_bg= PRIO_BG.get(prio, C_GRIS_CLR)

        vals = [
            ri - 4,
            PRIO_LBL.get(prio, f"P{prio}"),
            row["familia"],
            row["codigo"],
            row["descripcion"],
            row["umv"],
            row["precio"],
            row["precio_un"],
            row["motivo"],
        ]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci)
            c.border = border()

            if ci == 1:          # número
                fmt(c, v, bold=True, halign="center", bg=prio_bg)
            elif ci == 2:        # prioridad
                fmt(c, v, bold=True, bg=prio_bg, halign="center")
            elif ci in (7, 8):   # precios
                c.value = float(v) if v else 0
                c.font  = Font(name="Arial", size=9)
                c.number_format = "$#,##0"
                c.fill  = fill(row_bg)
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif ci == 9:        # motivo
                fmt(c, v, bg=row_bg, wrap=True)
            else:
                fmt(c, v, bg=row_bg)

        ws.row_dimensions[ri].height = 28

    # ── Nota al pie ──
    last_row = len(resultados) + 6
    ws.row_dimensions[last_row - 1].height = 6
    ws.merge_cells(f"A{last_row}:I{last_row}")
    c = ws.cell(row=last_row, column=1)
    c.value = "  🟢 P1 Core = insumo imprescindible (sin esto no operan)   🟡 P2 Alto = diferenciador de calidad   🔴 P3 Complemento = buena adición a la canasta"
    c.font  = Font(name="Arial", italic=True, size=8, color="595959")
    c.fill  = fill("F2F2F2")

    # ── Anchos de columna ──
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 52
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 11
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 52

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    wb.save(output_path)
    return output_path


# ── API pública ──────────────────────────────────────────────────────────────

def generar_top15(config: dict) -> str:
    """
    config = {
        "cliente": "Nombre del cliente",
        "tipo": "PASTELERÍA",
        "catalogo": "ruta/al/catalogo.xlsx",
        "output": "ruta/output.xlsx",
        "familias": [
            {
                "familia": "COBERTURA",
                "prioridad": "1",
                "consumo": "MUY ALTO",
                "motivo": "Chocolate para tortas y brownies",
                "codigo_preferido": "104551100"  # opcional
            },
            ...
        ]
    }
    """
    df = leer_catalogo(config["catalogo"])
    resultados = seleccionar_top15(df, config["familias"])
    return generar_excel(
        resultados,
        config["cliente"],
        config["tipo"],
        config["output"]
    )


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Genera Top 15 SKUs ICB para un cliente")
    parser.add_argument("--cliente",  required=True)
    parser.add_argument("--tipo",     required=True)
    parser.add_argument("--catalogo", required=True)
    parser.add_argument("--familias", required=True,
                        help="Familias separadas por coma, en orden de prioridad")
    parser.add_argument("--output",   required=True)
    args = parser.parse_args()

    df = leer_catalogo(args.catalogo)
    familias_list = [f.strip() for f in args.familias.split(",")]

    # Mapeo automático de prioridades: primeros 40% = P1, siguientes 40% = P2, resto = P3
    n = len(familias_list)
    fam_config = []
    for i, fam in enumerate(familias_list):
        if i < n * 0.4:
            prio = "1"
        elif i < n * 0.8:
            prio = "2"
        else:
            prio = "3"
        fam_config.append({
            "familia": fam,
            "prioridad": prio,
            "consumo": "ALTO",
            "motivo": f"Insumo relevante para {args.tipo.lower()}"
        })

    resultados = seleccionar_top15(df, fam_config)
    out = generar_excel(resultados, args.cliente, args.tipo, args.output)
    print(f"✓ Excel generado: {out}")
    print(f"  {len(resultados)} SKUs incluidos")
    for r in resultados:
        print(f"  [{r['prioridad']}] {r['familia']:<28} {r['codigo']} — {r['descripcion'][:50]}")


if __name__ == "__main__":
    main()
